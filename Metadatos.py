import pandas as pd 
from openpyxl import Workbook
import os
import json

class ETL:
    def __init__(self, path_carpeta_entrada,
                 path_carpeta_salida,
                 nombre_archivo_excel,
                 nombre_archivo_json,
                 tipos_de_datos_sql=None,
                 separador=",",
                 ):
        """Inicializamos el ETL."""

        self.path_carpeta_entrada = path_carpeta_entrada
        self.path_carpeta_salida = path_carpeta_salida
        self.nombre_archivo_excel = nombre_archivo_excel
        self.nombre_archivo_json = nombre_archivo_json
        self.separador = separador

        self.ruta_excel = os.path.join(self.path_carpeta_salida, self.nombre_archivo_excel)
        self.ruta_json = os.path.join(self.path_carpeta_salida, self.nombre_archivo_json)
        self.wb = Workbook()

        self.informacion_archivos = {}

        self.TIPOS_DE_DATOS_SQL = tipos_de_datos_sql or {
            'int64': 'INT',
            'float64': 'FLOAT',
            'object': 'VARCHAR',
            'datetime64': 'DATETIME',
            'bool': 'BIT'
        }

        self.crear_carpetas()
        
    def crear_carpetas(self):
        """Crea las carpetas de entrada y salida si no existen."""
        for carpeta in [self.path_carpeta_entrada, self.path_carpeta_salida]:
            if not os.path.exists(carpeta):
                os.makedirs(carpeta)
                print(f"Se creó la carpeta: {carpeta}")

    def cargar_csvs_en_carpeta(self, separador=","):
        """Cargamos archivos CSV en la carpeta de entrada."""
        dfs = {}  

        for archivo in os.listdir(self.path_carpeta_entrada):
            if archivo.endswith(".csv") or archivo.endswith(".txt"):
                archivo_path = os.path.join(self.path_carpeta_entrada, archivo)
                if archivo.startswith("MML"):
                    separador = ";"
                elif archivo.startswith("TAR_PROD"):
                    separador = ","
                elif archivo.startswith("ARG_UDO"):
                    separador = ","
                try:
                    # Leemos el archivo CSV y crear el DataFrame
                    df = pd.read_csv(archivo_path, sep= separador)
                    
                    # Agregamos el DataFrame al diccionario
                    dfs[archivo] = df
                except pd.errors.ParserError as e:
                    print(f"Error al cargar {archivo}: {e}")
        
        return dfs

    def write_excel(self, dataframes):
        """Escribimos los metadatos en un archivo Excel."""
        
        for archivo, df in dataframes.items():
            ws = self.wb.create_sheet(title=archivo)
            
            # Escribimos los encabezados
            ws.append(["Nombre del Campo", "Tipo de Campo", "Longitud del Campo", "Tiene Valores Vacíos", "Cantidad de Valores Vacíos", "Cantidad de Valores Únicos"])
            
            # Escribimos la metadata
            for columna, tipo in zip(df.columns, df.dtypes):
                tipo_sql = self.TIPOS_DE_DATOS_SQL.get(str(tipo), str(tipo))
                longitud = df[columna].apply(str).str.len().max()
                tiene_valores_vacios = df[columna].isnull().any()
                cantidad_valores_vacios = df[columna].isnull().sum()
                valores_unicos = df[columna].nunique()
                
               
                # Agregamos información a la hoja de Excel
                ws.append([columna, str(tipo_sql), longitud, tiene_valores_vacios, cantidad_valores_vacios, valores_unicos])

        # Eliminamos la hoja de inicio por defecto
        del self.wb[self.wb.sheetnames[0]]

        # Guardamos el libro de Excel
        self.wb.save(self.ruta_excel)
        print("Archivo de Excel guardado en:", self.ruta_excel)


    def write_json(self, dataframes):
        """Escribimos los metadatos en un archivo JSON."""

        for archivo, df in dataframes.items():
            for columna, tipo in zip(df.columns, df.dtypes):

                tipo_sql = self.TIPOS_DE_DATOS_SQL.get(str(tipo), str(tipo))
                longitud = df[columna].apply(str).str.len().max()
                tiene_valores_vacios = df[columna].isnull().any()
                cantidad_valores_vacios = df[columna].isnull().sum()
                valores_unicos = df[columna].nunique()

                # Convertimos int64 a tipos nativos de Python
                if isinstance(longitud, pd.Int64Dtype):
                    longitud = int(longitud)
                if isinstance(cantidad_valores_vacios, pd.Int64Dtype):
                    cantidad_valores_vacios = int(cantidad_valores_vacios)



                # Agregamos información al JSON
                if archivo not in self.informacion_archivos:
                    self.informacion_archivos[archivo] = {"campos": []}
                
                self.informacion_archivos[archivo]["campos"].append({
                    "nombre": columna,
                    "tipo": tipo_sql,
                    "longitud": longitud,
                    "tiene_valores_vacios": tiene_valores_vacios,
                    "cantidad_valores_vacios": cantidad_valores_vacios,
                    "valores_unicos": valores_unicos
                })
        
        with open(self.ruta_json, "w") as archivo_json:
            json.dump(self.informacion_archivos, archivo_json, indent=4, default=str)

        print("Información de metadatos guardada en:", self.ruta_json)

    def run(self):
        """Ejecutamos el proceso completo."""

        dataframes = self.cargar_csvs_en_carpeta()
        self.write_excel(dataframes)
        self.write_json(dataframes)


if __name__ == "__main__":

    path_carpeta_entrada = os.path.join(os.getcwd(), 'TEMP_DATA')
    path_carpeta_salida = os.path.join(os.getcwd(), 'TEMP_DATA', 'METADATOS')

    etl = ETL(
        path_carpeta_entrada,
        path_carpeta_salida,
        nombre_archivo_excel="Metadatos.xlsx",
        nombre_archivo_json="Metadatos.json"
    )
    etl.run()